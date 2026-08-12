# -*- coding: utf-8 -*-
"""Gerador v2: padronizacao PDAF23-26.xlsx com regras de expansao do usuario,
escolhas 'ok' preservadas e verificacao por coluna RA."""
import unicodedata
import re
import openpyxl
import json

XLSX = 'PDAF23-26.xlsx'
REPORT = 'relatorio_padronizacao.txt'

# ---------- normalizacao ----------
def norm(s):
    s = unicodedata.normalize('NFD', str(s or ''))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('\u2013', '-').replace('\u2014', '-')
    return ' '.join(s.upper().strip().split())

def canonical(s):
    s = norm(s)
    # sinonimos de prefixo (ordem importa: mais especifico primeiro)
    s = re.sub(r'\bESCOLA CLASSE\b', 'EC', s)
    s = re.sub(r'\bESCOLA PARQUE\b', 'EP', s)
    s = re.sub(r'\bESC\b', 'ESCOLA', s)
    s = re.sub(r'\bCENTRO DE ENSINO MEDIO INTEGRADO\b', 'CEMI', s)
    s = re.sub(r'\bCENTRO DE ENSINO MEDIO\b', 'CEM', s)
    s = re.sub(r'\bCENTRO DE EDUCACAO INFANTIL\b', 'CEI', s)
    s = re.sub(r'\bCENTRO DE ENSINO ESPECIAL\b', 'CEE', s)
    s = re.sub(r'\bCENTRO DE ENSINO FUNDAMENTAL\b', 'CEF', s)
    s = re.sub(r'\bCENTRO EDUCACIONAL\b', 'CED', s)
    s = re.sub(r'\bCOORDENACAO REGIONAL DE ENSINO\b', 'CRE', s)
    s = re.sub(r'\bCOORDENACAO REGIONAL\b', 'CRE', s)
    s = re.sub(r'\bPROFESSORA\b', 'PROF', s)
    s = re.sub(r'\bPROFESSOR\b', 'PROF', s)
    s = re.sub(r'\bPROF\.', 'PROF', s)
    s = re.sub(r'[^A-Z0-9\s]', ' ', s)   # remove - ( ) / , . etc
    return re.sub(r'\s+', ' ', s).strip()

# ---------- mapeamentos manuais (raw normalizado -> chave final) ----------
MANUAL_MAP = {
    'CEM TN': 'CEM TAGUATINGA NORTE',
    'CEMEIT - CENTRO DE ENSINO MEDIO EIT': 'CEM EIT',
    'CED 310 DE SANTA MARIA/CEMI SANTA MARIA': 'CED 310 DE SANTA MARIA',
    'CEF ZILDA ARNS': 'CEF DOUTORA ZILDA ARNS',
    'CEF NOSSA SENHORA FATIMA': 'CEF NOSSA SENHORA DE FATIMA',
    'ESC MENINOS E MENINAS DO PARQUE': 'ESCOLA MENINOS E MENINAS DO PARQUE',
    'CRE CEILANDIA (ESCOLA PARQUE ANISIO TEIXEIRA)': 'EP ANISIO TEIXEIRA',
    'COORDENACAO REGIONAL DE CEILANDIA (CED 17 DE CEILANDIA)': 'CED 17 DE CEILANDIA',
    'COORDENACAO REGIONAL DE ENSINO DE SAO SEBASTIAO (CED SAO BARTOLOMEU)': 'CED SAO BARTOLOMEU',
    'COORDENACAO REGIONAL DE ENSINO DE SAO SEBASTIAO (CED SAO FRANCISCO)': 'CED SAO FRANCISCO',
    'COORDENACAO REGIONAL DE ENSINO DE SAO SEBASTIAO (CEF DO BOSQUE)': 'CEF DO BOSQUE',
    'COORDENACAO REGIONAL DE ENSINO DE SAO SEBASTIAO (CEM 01)': 'CEM 01 DE SAO SEBASTIAO',
    'COORDENACAO REGIONAL DE TAGUATINGA (CEM AVE BRANCA - CEMAB)': 'CEM AVE BRANCA',
    'COORDENACAO REGIONAL DE TAGUATINGA (POLO DE ALTAS HABILIDADES/SUPERDOTACAO DE TAGUATINGA)': 'POLO DE ALTAS HABILIDADES SUPERDOTACAO DE TAGUATINGA',
    'COORDENACAO REGIONAL DE ENISNO DO GAMA': 'CRE GAMA',
    'COORDENACAO REGIONAL DE TAGUATINGA': 'CRE TAGUATINGA',
    'CENTRO EDUCACIONAL INCRA 09 DE CEILANDIA': 'CED INCRA 09',
    'CENTRO EDUCACIONAL ENGENHO DAS LAGES': 'CED ENGENHO DAS LAJES',
    'CENTRO EDUCACIONAL ESTELLA DOS CHERUBINS': 'CED STELLA DOS CHERUBINS GUIMARAES TROIS',
    'CENTRO DE ENSINO MEDIO 01': 'CEM 01 DE PLANALTINA',
    'CENTRO DE ENSINO FUNDAMENTAL 01': 'CEF 01 DE PLANALTINA',
    'ESCOLA CLASSE 65': 'EC 65 DE CEILANDIA',
    'ESCOLA CLASSE 66': 'EC 66 DE CEILANDIA',
    'ESCOLA CLASSE 68': 'EC 68 DE CEILANDIA',
}

# Correcoes de RA confirmadas pelo usuario (grupo -> RA padrao)
RA_FIX = {
    'CEF DOUTORA ZILDA ARNS': 'Itapoã',   # escola e no Itapoa (confirmado)
}

# ---------- carrega escolhas 'ok' do relatorio anotado anterior ----------
OK_CHOICES_FILE = 'escolhas_padrao.json'

def parse_ok_choices():
    # 1) sidecar JSON persistente (fonte de verdade)
    try:
        with open(OK_CHOICES_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        pass
    # 2) fallback: parseia relatorio (formato v1 "ok" ou v2 "<-- PADRAO")
    choices = {}
    try:
        with open(REPORT, encoding='utf-8') as f:
            txt = f.read()
        sec1 = txt.split('1) ESCOLAS COM GRAFIA DUPLICADA')[1].split('2) ENTRADAS QUE NAO SAO ESCOLA')[0]
        cur = None
        for line in sec1.splitlines():
            m = re.match(r'^\s+\[(.+)\]$', line)
            if m:
                cur = m.group(1)
                continue
            m2 = re.match(r'^\s+"(.*)"\s+->\s+abas:.*?(?:ok|<-- PADRAO)\s*$', line)
            if m2 and cur is not None:
                choices[cur] = m2.group(1)
    except Exception:
        pass
    return choices

OK_CHOICES = parse_ok_choices()
try:
    with open(OK_CHOICES_FILE, 'w', encoding='utf-8') as f:
        json.dump(OK_CHOICES, f, ensure_ascii=False, indent=1)
except Exception:
    pass

# ---------- leitura da planilha ----------
wb = openpyxl.load_workbook(XLSX, data_only=True)

rows = []  # (sheet, linha, raw, key, ra_norm)
for ws in wb.worksheets:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v is None:
            continue
        raw = ' '.join(str(v).strip().split())
        if not raw:
            continue
        ra = ws.cell(row=r, column=3).value
        key = MANUAL_MAP.get(norm(raw), canonical(raw))
        if ra is not None and key in RA_FIX:
            ra = RA_FIX[key]
        rows.append((ws.title, r, raw, key, norm(ra) if ra else ''))

# ---------- agrupamento ----------
groups = {}
for sheet, r, raw, key, ra in rows:
    groups.setdefault(key, []).append((sheet, r, raw, ra))

# ---------- verificacao por RA ----------
ra_conflito = {}
for key, membros in groups.items():
    ras = {ra for _, _, _, ra in membros}
    if len(ras) > 1:
        ra_conflito[key] = sorted(ras)

# ---------- entradas nao-escola / incompletas ----------
nao_escola_keys = sorted({key for key in groups
                          if key.startswith('CRE ') or key.startswith('COORDENACAO')})
sem_regiao = sorted({k for k in groups
                     if re.fullmatch(r'(EC|ESCOLA CLASSE|CEF|CED|CEM|CEMI|CEI|JI|CEE|CIL|CAIC)\s+\d+', k)})

# ---------- nome padrao (escolha ok ou chave canonica) ----------
def nome_padrao(key):
    return OK_CHOICES.get(key, key)

# ---------- seed com TODAS as colunas + Escola_Padrao ----------
import csv
cabeçalhos = []  # uniao de colunas na ordem de aparicao
col_por_aba = {}
for ws in wb.worksheets:
    hdr = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        hdr.append(v if v is not None else f'COL{c}')
    col_por_aba[ws.title] = hdr
    for h in hdr:
        if h not in cabeçalhos:
            cabeçalhos.append(h)

with open('seed_padronizado.csv', 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.writer(f, delimiter=';')
    w.writerow(cabeçalhos + ['Escola_Padrao'])
    for ws in wb.worksheets:
        hdr = col_por_aba[ws.title]
        for r in range(2, ws.max_row + 1):
            raw_row = {h: ws.cell(row=r, column=c + 1).value for c, h in enumerate(hdr)}
            v = raw_row.get('Escola')
            if v is None:
                continue
            raw = ' '.join(str(v).strip().split())
            if not raw:
                continue
            key = MANUAL_MAP.get(norm(raw), canonical(raw))
            linha = [raw_row.get(h, '') for h in cabeçalhos]
            if key in RA_FIX and raw_row.get('RA') is not None:
                linha[cabeçalhos.index('RA')] = RA_FIX[key]
            # Ano nao numerico (ex.: numero da OB digitado na coluna) -> ano da aba
            try:
                float(str(linha[cabeçalhos.index('Ano')]).strip())
            except (TypeError, ValueError):
                linha[cabeçalhos.index('Ano')] = ws.title
            linha.append(nome_padrao(key))
            w.writerow(linha)

# ---------- relatorio ----------
rep = []
A = rep.append
A('=' * 78)
A('RELATORIO DE PADRONIZACAO v2 - PDAF23-26.xlsx')
A('=' * 78)
A('')
A(f'Abas: {wb.sheetnames}')
A(f'Total de linhas: {len(rows)}')
A(f'Chaves finais (unidades unicas): {len(groups)}')
A(f'Grupos com grafia duplicada: {len([k for k, v in groups.items() if len({raw for _, _, raw, _ in v}) > 1])}')
A('')

A('-' * 78)
A('1) ESCOLAS COM GRAFIA DUPLICADA  (PADRAO = sua escolha "ok")')
A('-' * 78)
for key in sorted(groups):
    membros = groups[key]
    raws = sorted({raw for _, _, raw, _ in membros})
    if len(raws) <= 1:
        continue
    A(f'')
    A(f'  [{key}]')
    padrao = OK_CHOICES.get(key, key)
    from collections import defaultdict
    abas_por_raw = defaultdict(set)
    for s, _, raw_, _ in membros:
        abas_por_raw[raw_].add(s)
    for raw in raws:
        marca = '  <-- PADRAO' if raw == padrao else ''
        A(f'      "{raw}"  -> abas: {", ".join(sorted(abas_por_raw[raw]))}{marca}')
    A(f'      PADRAO: "{padrao}"')

A('')
A('-' * 78)
A('2) ENTRADAS QUE NAO SAO ESCOLA (CRE / Coordenacao Regional)')
A('-' * 78)
for key in nao_escola_keys:
    A(f'  {key}  ({len(groups[key])} linha(s))')

A('')
A('-' * 78)
A('3) NOMES INCOMPLETOS / SEM REGIAO')
A('-' * 78)
if sem_regiao:
    for key in sem_regiao:
        A(f'  {key}')
        for sheet, r, raw, ra in groups[key]:
            A(f'      "{raw}" ({sheet} linha {r}, RA={ra})')
else:
    A('  (nenhum - todos os nomes incompletos foram resolvidos)')

A('')
A('-' * 78)
A('4) VERIFICACAO POR RA (grupos com RA divergente entre linhas)')
A('-' * 78)
if ra_conflito:
    A('ATENCAO: estes grupos tem linhas com RA diferentes - conferir:')
    for key in sorted(ra_conflito):
        A(f'')
        A(f'  [{key}]  RAs: {", ".join(ra_conflito[key])}')
        for sheet, r, raw, ra in groups[key]:
            A(f'      "{raw}" ({sheet} linha {r}) -> RA={ra}')
else:
    A('Nenhum grupo com RA divergente. Todos os agrupamentos batem com a regional.')

A('')
A('-' * 78)
A('5) NOMES COM FORMA EXTENSA CONVERTIDOS PELAS NOVAS REGRAS')
A('-' * 78)
for sheet, r, raw, key, ra in rows:
    up = raw.upper()
    if any(p in up for p in ['CENTRO DE ENSINO', 'CENTRO EDUCACIONAL', 'COORDENACAO']):
        A(f'  "{raw}" ({sheet} linha {r}) -> {key}')

A('')
A('-' * 78)
A('6) MAPPING DICT PARA EMENDAS.PY')
A('-' * 78)
A('PADRONIZAR_ESCOLA = {')
for raw_n in sorted(MANUAL_MAP):
    A(f'    {json.dumps(raw_n)}: {json.dumps(MANUAL_MAP[raw_n])},')
A('}')
A('')

A('-' * 78)
A('7) LISTA FINAL - chave canonica | NOME PADRAO (para o seed)')
A('-' * 78)
for key in sorted(groups):
    padrao = OK_CHOICES.get(key, key)
    A(f'  {key}  |  {padrao}')

out = '\n'.join(rep)
with open(REPORT, 'w', encoding='utf-8') as f:
    f.write(out)
print(out)
